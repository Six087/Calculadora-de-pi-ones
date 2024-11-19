from ctypes import alignment
import tkinter as tk
from openpyxl import Workbook

def guardar_numeros_en_excel():
    # Crear un nuevo archivo Excel
    workbook = Workbook()
    hoja = workbook.active
    hoja.title = "Resultados de la Calculadora"  # Nombre de la hoja

    # Obtener los valores de las variables StringVar y escribirlos en la hoja de Excel
    hoja.column_dimensions['A'].width = 30
    hoja.column_dimensions['B'].width = 30
    hoja.column_dimensions['C'].width = 30  # Establece el ancho de la columna A
    hoja.cell(row=1, column=1, value="DIAMETRO/RADIO DE UN PIÑON")
    hoja.cell(row=2, column=1, value=radio_result.get())
    hoja.cell(row=3, column=1, value=diametro_result.get())

    hoja.cell(row=1, column=2, value="DIAMETRO DE PIÑON")
    hoja.cell(row=4, column=1, value=diametro_pinon_result.get())

    hoja.cell(row=1, column=2, value="CALCULADORA ASA")
    hoja.cell(row=5, column=1, value=asa_result.get())

    hoja.cell(row=1, column=2, value="CALCULADORA ASA")
    hoja.cell(row=6, column=1, value=diametro_ext_result.get())

    hoja.cell(row=1, column=2, value="CALCULADORA ASA")
    hoja.cell(row=7, column=1, value=paso_result.get())

    # Guardar el archivo Excel
    workbook.save("Resultados de la Calculadora.xlsx")

#Aca estan los calculos

def calcular_radio_diametro():
    perimetro = float(perimetro_entry.get())
    radio = perimetro / (2 * 3.14159265359)
    diametro = radio * 2
    radio_result.set(round(radio, 2))
    diametro_result.set(round(diametro, 2))

def calcular_diametro_pinon():
    cantidad_dientes = float(cantidad_dientes_entry.get())
    asa = float(asa_entry.get())
    diametro = (cantidad_dientes * asa) / 3.14
    diametro_pinon_result.set(round(diametro, 2))


def calcular_asa():
    diametro = float(diametro_entry.get())
    dientes = float(dientes_entry.get())
    asa = (diametro / dientes) * 3.14
    asa_result.set(round(asa, 2))


def calcular_diametro_ext():
    cantidad_dientes = float(cantidad_dientes_ext_entry.get())
    modulo = float(modulo_entry.get())
    diametro = (cantidad_dientes + 2) * modulo
    diametro_ext_result.set(round(diametro, 2))


def calcular_paso():
    diametro = float(diametro_paso_entry.get())
    cantidad_dientes = float(cantidad_dientes_paso_entry.get())
    modulo = diametro / (cantidad_dientes + 2)
    paso_result.set(round(modulo, 2))

#Resolucion y cosas de la ventana del progrma

ventana = tk.Tk()
ventana.title("Calculadora de Piñones")
ventana.geometry("1280x840")
ventana.configure(bg="black")

#tamaños de intefaz
 
x_left = 0
ancho_seccion = 200
alto_seccion = 30

# primera calculadora 1

seccion1_label = tk.Label(ventana, text="DIAMETRO/RADIO DE UN PIÑON")
seccion1_label.pack()
seccion1_label.configure(bg="grey", fg="white")
seccion1_label.place(x=x_left, y=0)
seccion1_label.place(x=5, y=5, width=ancho_seccion, height=alto_seccion)

Numero_de_dientes = tk.Label(ventana, text="Numero de dientes (Z):")
Numero_de_dientes.pack()
Numero_de_dientes.configure(bg="grey", fg="white")
Numero_de_dientes.place(
    x=x_left+5, y=36, width=ancho_seccion, height=alto_seccion)
Numero_de_dientes_entry = tk.Entry(ventana)
Numero_de_dientes_entry.pack()
Numero_de_dientes_entry.configure(bg="grey", fg="white")
Numero_de_dientes_entry.place(
    x=x_left+5, y=66, width=ancho_seccion, height=alto_seccion)

Ancho_de_diente_label = tk.Label(ventana, text="Ancho de diente (d):")
Ancho_de_diente_label.pack()
Ancho_de_diente_label.configure(bg="grey", fg="white")
Ancho_de_diente_label.place(
    x=x_left+5, y=96, width=ancho_seccion, height=alto_seccion)
Ancho_de_diente_entry = tk.Entry(ventana)
Ancho_de_diente_entry.pack()
Ancho_de_diente_entry.configure(bg="grey", fg="white")
Ancho_de_diente_entry.place(
    x=x_left+5, y=126, width=ancho_seccion, height=alto_seccion)

Paso_Label = tk.Label(ventana, text="Paso (p):")
Paso_Label.pack()
Paso_Label.configure(bg="grey", fg="white")
Paso_Label.place(x=x_left+5, y=156, width=ancho_seccion, height=alto_seccion)
Paso_Entry = tk.Entry(ventana)
Paso_Entry.pack()
Paso_Entry.configure(bg="grey", fg="white")
Paso_Entry.place(x=x_left+5, y=186, width=ancho_seccion, height=alto_seccion)

perimetro_label = tk.Label(ventana, text="Perimetro (P):")
perimetro_label.pack()
perimetro_label.configure(bg="grey", fg="white")
perimetro_label.place(
    x=x_left+5, y=216, width=ancho_seccion, height=alto_seccion)
perimetro_entry = tk.Entry(ventana)
perimetro_entry.pack()
perimetro_entry.configure(bg="grey", fg="white")
perimetro_entry.place(
    x=x_left+5, y=246, width=ancho_seccion, height=alto_seccion)

calcular_radio_diametro_button = tk.Button(
    ventana, text="Calcular", command=calcular_radio_diametro)
calcular_radio_diametro_button.pack()
calcular_radio_diametro_button.configure(bg="grey", fg="white")
calcular_radio_diametro_button.place(
    x=x_left+5, y=276, width=ancho_seccion, height=alto_seccion)


radio_label = tk.Label(ventana, text="Radio:")
radio_result = tk.StringVar()
radio_label.pack()
radio_label.configure(bg="grey", fg="white")
radio_label.place(x=x_left+5, y=306, width=ancho_seccion, height=alto_seccion)
radio_result_label = tk.Label(ventana, textvariable=radio_result)
radio_result_label.pack()
radio_result_label.configure(bg="grey", fg="white")
radio_result_label.place(x=x_left+170, y=306, width=35, height=30)

diametro_label = tk.Label(ventana, text="Diametro:")
diametro_label.pack()
diametro_label.configure(bg="grey", fg="white")
diametro_label.place(
    x=x_left+5, y=336, width=ancho_seccion, height=alto_seccion)
diametro_result = tk.StringVar()
diametro_result_label = tk.Label(ventana, textvariable=diametro_result)
diametro_result_label.pack()
diametro_result_label.configure(bg="grey", fg="white")
diametro_result_label.place(x=x_left+170, y=336, width=35, height=30)

#########################################################################################################

#Segunda calculadora

seccion2_label = tk.Label(ventana, text="DIAMETRO DE PIÑON")
seccion2_label.pack()
seccion2_label.configure(bg="grey", fg="white")
seccion2_label.place(x=x_left+210, y=5, width=ancho_seccion, height=alto_seccion)

cantidad_dientes_label = tk.Label(ventana, text="Cantidad de dientes:")
cantidad_dientes_label.pack()
cantidad_dientes_label.configure(bg="grey", fg="white")
cantidad_dientes_label.place(x=x_left+210, y=35, width=ancho_seccion, height=alto_seccion)
cantidad_dientes_entry = tk.Entry(ventana)
cantidad_dientes_entry.pack()
cantidad_dientes_entry.configure(bg="grey", fg="white")
cantidad_dientes_entry.place(x=x_left+210, y=65, width=ancho_seccion, height=alto_seccion)

asa_label = tk.Label(ventana, text="ASA:")
asa_label.pack()
asa_label.configure(bg="grey", fg="white")
asa_label.place(x=x_left+210, y=95, width=ancho_seccion, height=alto_seccion)
asa_entry = tk.Entry(ventana)
asa_entry.pack()
asa_entry.configure(bg="grey", fg="white")
asa_entry.place(x=x_left+210, y=125, width=ancho_seccion, height=alto_seccion)

calcular_diametro_pinon_button = tk.Button(
    ventana, text="Calcular", command=calcular_diametro_pinon)
calcular_diametro_pinon_button.pack()
calcular_diametro_pinon_button.configure(bg="grey", fg="white")
calcular_diametro_pinon_button.place(x=x_left+210, y=155, width=ancho_seccion, height=alto_seccion)

diametro_pinon_result = tk.StringVar()

diametro_pinon_label = tk.Label(ventana, text="Diametro:")
diametro_pinon_label.pack()
diametro_pinon_label.configure(bg="grey", fg="white")
diametro_pinon_label.place(x=x_left+210, y=185, width=ancho_seccion, height=alto_seccion)
diametro_pinon_result_label = tk.Label(
    ventana, textvariable=diametro_pinon_result)
diametro_pinon_result_label.pack()
diametro_pinon_result_label.configure(bg="grey", fg="white")
diametro_pinon_result_label.place(x=x_left+375, y=185, width=35, height=30)
############################################################################################################

#Tercera calculadora

seccion3_label = tk.Label(ventana, text="CALCULADORA ASA")
seccion3_label.pack()
seccion3_label.configure(bg="grey", fg="white")
seccion3_label.place(x=x_left+415, y=5, width=ancho_seccion, height=alto_seccion)

diametro_label = tk.Label(ventana, text="Diametro:")
diametro_label.pack()
diametro_label.configure(bg="grey", fg="white")
diametro_label.place(x=x_left+415, y=35, width=ancho_seccion, height=alto_seccion)
diametro_entry = tk.Entry(ventana)
diametro_entry.pack()
diametro_entry.configure(bg="grey", fg="white")
diametro_entry.place(x=x_left+415, y=65, width=ancho_seccion, height=alto_seccion)

dientes_label = tk.Label(ventana, text="Dientes:")
dientes_label.pack()
dientes_label.configure(bg="grey", fg="white")
dientes_label.place(x=x_left+415, y=95, width=ancho_seccion, height=alto_seccion)
dientes_entry = tk.Entry(ventana)
dientes_entry.pack()
dientes_entry.configure(bg="grey", fg="white")
dientes_entry.place(x=x_left+415, y=125, width=ancho_seccion, height=alto_seccion)

calcular_asa_button = tk.Button(ventana, text="Calcular", command=calcular_asa)
calcular_asa_button.pack()
calcular_asa_button.configure(bg="grey", fg="white")
calcular_asa_button.place(x=x_left+415, y=155, width=ancho_seccion, height=alto_seccion)

asa_result = tk.StringVar()

asa_label = tk.Label(ventana, text="ASA:")
asa_label.pack()
asa_label.configure(bg="grey", fg="white")
asa_label.place(x=x_left+415, y=185, width=ancho_seccion, height=alto_seccion)
asa_result_label = tk.Label(ventana, textvariable=asa_result)
asa_result_label.pack()
asa_result_label.configure(bg="grey", fg="white")
asa_result_label.place(x=x_left+580, y=185, width=35, height=30)
#####################################################################################################################################

#Cuarta Calculadora

seccion4_label = tk.Label(ventana, text="CALCULAR EL Ø EXT")
seccion4_label.pack()
seccion4_label.configure(bg="grey", fg="white")
seccion4_label.place(x=x_left+620, y=5, width=ancho_seccion, height=alto_seccion)

cantidad_dientes_ext_label = tk.Label(ventana, text="Cantidad de dientes:")
cantidad_dientes_ext_label.pack()
cantidad_dientes_ext_label.configure(bg="grey", fg="white")
cantidad_dientes_ext_label.place(x=x_left+620, y=35, width=ancho_seccion, height=alto_seccion)
cantidad_dientes_ext_entry = tk.Entry(ventana)
cantidad_dientes_ext_entry.pack()
cantidad_dientes_ext_entry.configure(bg="grey", fg="white")
cantidad_dientes_ext_entry.place(x=x_left+620, y=65, width=ancho_seccion, height=alto_seccion)

modulo_label = tk.Label(ventana, text="Modulo:")
modulo_label.pack()
modulo_label.configure(bg="grey", fg="white")
modulo_label.place(x=x_left+620, y=95, width=ancho_seccion, height=alto_seccion)
modulo_entry = tk.Entry(ventana)
modulo_entry.pack()
modulo_entry.configure(bg="grey", fg="white")
modulo_entry.place(x=x_left+620, y=125, width=ancho_seccion, height=alto_seccion)

calcular_diametro_ext_button = tk.Button(
    ventana, text="Calcular", command=calcular_diametro_ext)
calcular_diametro_ext_button.pack()
calcular_diametro_ext_button.configure(bg="grey", fg="white")
calcular_diametro_ext_button.place(x=x_left+620, y=155, width=ancho_seccion, height=alto_seccion)

diametro_ext_result = tk.StringVar()

diametro_ext_label = tk.Label(ventana, text="Diametro:")
diametro_ext_label.pack()
diametro_ext_label.configure(bg="grey", fg="white")
diametro_ext_label.place(x=x_left+620, y=185, width=ancho_seccion, height=alto_seccion)
diametro_ext_result_label = tk.Label(ventana, textvariable=diametro_ext_result)
diametro_ext_result_label.pack()
diametro_ext_result_label.configure(bg="grey", fg="white")
diametro_ext_result_label.place(x=x_left+785, y=185, width=35 , height=30)
########################################################################################################################################

#Quinta Calculadora

seccion5_label = tk.Label(ventana, text="CALCULAR EL PASO")
seccion5_label.pack()
seccion5_label.configure(bg="grey", fg="white")
seccion5_label.place(x=x_left+825, y=5, width=ancho_seccion, height=alto_seccion)

diametro_paso_label = tk.Label(ventana, text="Diametro:")
diametro_paso_label.pack()
diametro_paso_label.configure(bg="grey", fg="white")
diametro_paso_label.place(x=x_left+825, y=35, width=ancho_seccion, height=alto_seccion)
diametro_paso_entry = tk.Entry(ventana)
diametro_paso_entry.pack()
diametro_paso_entry.configure(bg="grey", fg="white")
diametro_paso_entry.place(x=x_left+825, y=65, width=ancho_seccion, height=alto_seccion)

cantidad_dientes_paso_label = tk.Label(ventana, text="Cantidad de dientes:")
cantidad_dientes_paso_label.pack()
cantidad_dientes_paso_label.configure(bg="grey", fg="white")
cantidad_dientes_paso_label.place(x=x_left+825, y=95, width=ancho_seccion, height=alto_seccion)
cantidad_dientes_paso_entry = tk.Entry(ventana)
cantidad_dientes_paso_entry.pack()
cantidad_dientes_paso_entry.configure(bg="grey", fg="white")
cantidad_dientes_paso_entry.place(x=x_left+825, y=125,width=ancho_seccion, height=alto_seccion)

calcular_paso_button = tk.Button(
    ventana, text="Calcular", command=calcular_paso)
calcular_paso_button.pack()
calcular_paso_button.configure(bg="grey", fg="white")
calcular_paso_button.place(x=x_left+825, y=155, width=ancho_seccion, height=alto_seccion)

paso_result = tk.StringVar()

paso_label = tk.Label(ventana, text="Modulo:")
paso_label.pack()
paso_label.configure(bg="grey", fg="white")
paso_label.place(x=x_left+825, y=185, width=ancho_seccion, height=alto_seccion)
paso_result_label = tk.Label(ventana, textvariable=paso_result)
paso_result_label.pack()
paso_result_label.configure(bg="grey", fg="white")
paso_result_label.place(x=x_left+990, y=185, width=35 ,height=30)
###########################################################################################################################################

# Botón para guardar los números en Excel

boton_guardar_excel = tk.Button(
    ventana, text="Guardar Números en Excel", command=guardar_numeros_en_excel)
boton_guardar_excel.pack()
boton_guardar_excel.configure(bg="grey", fg="white")
boton_guardar_excel.place(x=x_left+1080, y=800, width=ancho_seccion, height=alto_seccion)


ventana.mainloop()

#Tareas pendientes 
#areglar los guarados del exel y que cuando se guarden esten ordenados
#crear marco de referencias ASA(ASA,Pulgadas,Medida)(no se tiene que cambiar nada solo de referencia de datos)
#crear calculadora de Potencia Bomba y de Polea Bomba)
