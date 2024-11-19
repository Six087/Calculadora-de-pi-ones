import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import Workbook
from datetime import datetime
import math
import unicodedata

class CalculadoraPinones:
    def __init__(self):
        # Configuración de la ventana principal
        self.ventana = tk.Tk()
        self.ventana.title("Calculadora de Piñones Industrial")
        self.ventana.geometry("1280x840")
        self.ventana.configure(bg='#f0f0f0')
        
        # Configuración del grid
        self.ventana.grid_rowconfigure(0, weight=1)
        self.ventana.grid_rowconfigure(1, weight=1)
        for i in range(3):
            self.ventana.grid_columnconfigure(i, weight=1)
        
        self.resultados = {}
        self.crear_variables()
        self.crear_barra_herramientas()
        self.crear_secciones()
        self.configurar_sincronizacion()

    def configurar_sincronizacion(self):
        """Configura la sincronización entre campos relacionados"""
        # Mapeo de campos que deben sincronizarse
        self.campos_sincronizados = {
            'dientes': [
                self.vars['radio_diametro_pinon']['numero_de_dientes_z'],
                self.vars['diametro_pinon']['dientes'],
#                self.vars['asa']['dientes'],
#                self.vars['diametro_ext']['dientes'],
#                self.vars['paso']['dientes'],
#                self.vars['calculo_engranaje']['numero_de_dientes_n']
            ],
#           'diametro': [
#                self.vars['asa']['diametro'],
#                self.vars['paso']['diametro']
#            ]
       }

        # Configura los trace para cada grupo de campos sincronizados
        for grupo in self.campos_sincronizados.values():
            for var in grupo:
                var.trace_add('write', lambda *args, grupo=grupo, var_actual=var: 
                    self.sincronizar_campos(grupo, var_actual))

    def sincronizar_campos(self, grupo, var_actual):
        """Sincroniza los campos relacionados"""
        try:
            valor = var_actual.get()
            if valor:  # Solo sincroniza si hay un valor
                for var in grupo:
                    if var != var_actual:  # Evita recursión infinita
                        var.set(valor)
        except tk.TclError:
            pass  # Ignora errores de actualización de variables

    def normalizar_texto(self, texto):
        """Normaliza el texto eliminando acentos y caracteres especiales"""
        texto = ''.join(c for c in unicodedata.normalize('NFD', texto)
                       if unicodedata.category(c) != 'Mn')
        return texto.lower().replace('/', '_').replace(' ', '_').replace('(', '').replace(')', '')

    def crear_variables(self):
        """Inicializa las variables para cada sección de la calculadora"""
        # Variables independientes para cada sección
        self.vars = {
            'radio_diametro_pinon': {
                'numero_de_dientes_z': tk.StringVar(),
                'ancho_del_diente_d': tk.StringVar(),
                'paso_p': tk.StringVar(),
                'perimetro': tk.StringVar(),
                'radio': tk.StringVar(),
                'diametro': tk.StringVar(),
                'resultado': tk.StringVar()
            },
            'diametro_pinon': {
                'dientes': tk.StringVar(),
                'asa': tk.StringVar(),
                'resultado': tk.StringVar()
            },
            'asa': {
                'diametro': tk.StringVar(),
                'dientes': tk.StringVar(),
                'resultado': tk.StringVar()
            },
            'diametro_ext': {
                'dientes': tk.StringVar(),
                'modulo': tk.StringVar(),
                'resultado': tk.StringVar()
            },
            'paso': {
                'diametro': tk.StringVar(),
                'dientes': tk.StringVar(),
                'resultado': tk.StringVar()
            },
            'calculo_engranaje': {
                'numero_de_dientes_n': tk.StringVar(),
                'paso_diametral_pd': tk.StringVar(),
                'diametro_primitivo': tk.StringVar(),
                'paso_circular': tk.StringVar(),
                'modulo': tk.StringVar(),
                'resultado': tk.StringVar()
            }
        }

    def crear_barra_herramientas(self):
        """Crea la barra de herramientas con botones de acción"""
        barra_herramientas = tk.Frame(self.ventana, bg='#e0e0e0')
        barra_herramientas.grid(row=2, column=0, columnspan=3, sticky='ew', padx=10, pady=5)

        btn_referencias = tk.Button(barra_herramientas, 
                                  text="Referencias ASA", 
                                  command=self.mostrar_referencias,
                                  bg='#4CAF50', 
                                  fg='white')
        btn_referencias.pack(side=tk.LEFT, padx=5)

        btn_excel = tk.Button(barra_herramientas, 
                            text="Guardar en Excel", 
                            command=self.guardar_excel,
                            bg='#4CAF50', 
                            fg='white')
        btn_excel.pack(side=tk.LEFT, padx=5)

        btn_limpiar = tk.Button(barra_herramientas, 
                               text="Limpiar Campos", 
                               command=self.limpiar_campos,
                               bg='#4CAF50', 
                               fg='white')
        btn_limpiar.pack(side=tk.LEFT, padx=5)

    def limpiar_campos(self):
        """Limpia todos los campos de entrada"""
        for seccion in self.vars.values():
            for var in seccion.values():
                var.set('')

    def crear_secciones(self):
        """Crea las diferentes secciones de la calculadora"""
        calculadoras = [
            ("Radio/Diámetro Piñón", ["Número de Dientes (Z)", "Ancho del Diente (d)", "Paso (p)", "Perímetro"], self.calcular_radio_diametro),
            ("Diámetro Piñón", ["Dientes", "ASA"], self.calcular_diametro_pinon),
            ("ASA", ["Diámetro", "Dientes"], self.calcular_asa),
            ("Diámetro Ext", ["Dientes", "Módulo"], self.calcular_diametro_ext),
            ("Paso", ["Diámetro", "Dientes"], self.calcular_paso),
            ("Cálculo Engranaje", ["Número de Dientes (N)", "Paso Diametral (Pd)"], self.calcular_engranaje)
        ]
        
        for i, (titulo, campos, funcion) in enumerate(calculadoras):
            row = i // 3
            col = i % 3
            self.crear_seccion(titulo, campos, funcion, row, col)

    def crear_seccion(self, titulo, campos, funcion_calculo, fila, columna):
        """Crea una sección individual de la calculadora"""
        frame = tk.Frame(self.ventana, bd=2, relief='groove', bg='#e0e0e0')
        frame.grid(row=fila, column=columna, padx=10, pady=10, sticky='nsew')
        
        titulo_label = tk.Label(frame, text=titulo, font=('Arial', 12, 'bold'), bg='#e0e0e0')
        titulo_label.grid(row=0, column=0, columnspan=2, pady=5)
        
        entries = {}
        seccion_key = self.normalizar_texto(titulo)
        
        for i, campo in enumerate(campos, start=1):
            tk.Label(frame, text=f"{campo}:", bg='#e0e0e0').grid(row=i, column=0, padx=5, pady=2, sticky='e')
            campo_normalizado = self.normalizar_texto(campo)
            
            var = self.vars[seccion_key].get(campo_normalizado, tk.StringVar())
            entry = tk.Entry(frame, textvariable=var)
            entry.grid(row=i, column=1, padx=5, pady=2, sticky='ew')
            entries[campo] = entry
        
        tk.Button(frame, 
                 text="Calcular", 
                 command=lambda: self.calcular_con_validacion(funcion_calculo, entries),
                 bg='#4CAF50', 
                 fg='white').grid(row=len(campos)+1, column=0, columnspan=2, pady=10)
        
        # Personalizar la etiqueta del resultado según la sección
        resultado_texto = self.obtener_texto_resultado(titulo)
        tk.Label(frame, text=resultado_texto, bg='#e0e0e0').grid(row=len(campos)+2, column=0, padx=5, sticky='e')
        resultado_label = tk.Label(frame, textvariable=self.vars[seccion_key]['resultado'],
                                 bg='#e0e0e0', wraplength=200)
        resultado_label.grid(row=len(campos)+2, column=1, padx=5, sticky='w')

    def obtener_texto_resultado(self, titulo):
        """Devuelve el texto apropiado para la etiqueta de resultado según la sección"""
        if titulo == "Radio/Diámetro Piñón":
            return "Radio/Diámetro:"
        elif titulo == "Diámetro Piñón":
            return "Diámetro:"
        elif titulo == "ASA":
            return "ASA:"
        elif titulo == "Diámetro Ext":
            return "Diámetro:"
        elif titulo == "Paso":
            return "Módulo:"
        elif titulo == "Cálculo Engranaje":
            return "Resultados:"
        return "Resultado:"

    def calcular_con_validacion(self, funcion, entries):
        """Ejecuta una función de cálculo con validación de errores"""
        try:
            funcion(entries)
        except ValueError:
            messagebox.showerror("Error", "Por favor ingrese valores numéricos válidos")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def calcular_radio_diametro(self, entries):
        """Calcula el radio y diámetro del piñón"""
        try:
            num_dientes = float(entries["Número de Dientes (Z)"].get())
            ancho_diente = float(entries["Ancho del Diente (d)"].get())
            paso = float(entries["Paso (p)"].get())
            perimetro = float(entries["Perímetro"].get())
            
            PI = 3.14159265359
            radio = perimetro / (2 * PI)
            diametro = radio * 2
            
            resultado = f"Radio: {radio:.2f} mm\nDiámetro: {diametro:.2f} mm"
            self.vars['radio_diametro_pinon']['resultado'].set(resultado)
            
        except ValueError:
            messagebox.showerror("Error", "Por favor ingrese valores numéricos válidos")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def calcular_diametro_pinon(self, entries):
        """Calcula el diámetro del piñón"""
        try:
            dientes = float(entries["Dientes"].get())
            asa = float(entries["ASA"].get())
            diametro = (dientes * asa) / math.pi
            self.vars['diametro_pinon']['resultado'].set(f"{diametro:.2f} mm")
        except ValueError:
            messagebox.showerror("Error", "Por favor ingrese valores numéricos válidos")

    def calcular_asa(self, entries):
        """Calcula el valor ASA"""
        try:
            diametro = float(entries["Diámetro"].get())
            dientes = float(entries["Dientes"].get())
            asa = (diametro / dientes) * math.pi
            self.vars['asa']['resultado'].set(f"{asa:.2f}")
        except ValueError:
            messagebox.showerror("Error", "Por favor ingrese valores numéricos válidos")

    def calcular_diametro_ext(self, entries):
        """Calcula el diámetro exterior"""
        try:
            dientes = float(entries["Dientes"].get())
            modulo = float(entries["Módulo"].get())
            diametro = (dientes + 2) * modulo
            self.vars['diametro_ext']['resultado'].set(f"{diametro:.2f} mm")
        except ValueError:
            messagebox.showerror("Error", "Por favor ingrese valores numéricos válidos")

    def calcular_paso(self, entries):
        """Calcula el paso"""
        try:
            diametro = float(entries["Diámetro"].get())
            dientes = float(entries["Dientes"].get())
            modulo = diametro / (dientes + 2)
            self.vars['paso']['resultado'].set(f"{modulo:.2f}")
        except ValueError:
            messagebox.showerror("Error", "Por favor ingrese valores numéricos válidos")

    def calcular_engranaje(self, entries):
        """Calcula los parámetros del engranaje"""
        try:
            N = float(entries["Número de Dientes (N)"].get())
            Pd = float(entries["Paso Diametral (Pd)"].get())
            
            Dp = N / Pd
            Pc = (math.pi * Dp) / N
            modulo = 25.4 / Pd
            Dp_mm = Dp * 25.4
            
            resultado = (f"Diámetro primitivo: {Dp:.3f} in ({Dp_mm:.2f} mm)\n"
                        f"Paso circular: {Pc:.3f} in\n"
                        f"Módulo: {modulo:.3f} mm")
            
            self.vars['calculo_engranaje']['resultado'].set(resultado)
            
        except ValueError:
            messagebox.showerror("Error", "Por favor ingrese valores numéricos válidos")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def mostrar_referencias(self):
        """Muestra la ventana de referencias ASA"""
        ventana_ref = tk.Toplevel(self.ventana)
        ventana_ref.title("Referencias ASA")
        ventana_ref.geometry("600x600")
        ventana_ref.configure(bg='#f0f0f0')
        
        tree = ttk.Treeview(ventana_ref, columns=("ASA", "Pulgadas", "Medidas"), show="headings")
        tree.heading("ASA", text="ASA")
        tree.heading("Pulgadas", text="Pulgadas")
        tree.heading("Medidas", text="Medidas")
        
        referencias = [
            ("40", "1/2", "12.7"),
            ("50", "5/8", "15.875"),
            ("60", "3/4", "19.05"),
            ("80", "1", "25.4"),
            ("100", "1 1/4", "31.75"),
            ("120", "1 1/2", "38.1"),
        ]
        
        for ref in referencias:
            tree.insert("", "end", values=ref)
        
        tree.pack(fill='both', expand=True, padx=10, pady=10)

    def guardar_excel(self):
        """Guarda los resultados en un archivo Excel"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Resultados Calculadora"
            
            ws['A1'] = "Fecha y Hora"
            ws['B1'] = "Tipo de Cálculo"
            ws['C1'] = "Resultado"
            
            ws.column_dimensions['A'].width = 23
            ws.column_dimensions['B'].width = 44
            ws.column_dimensions['C'].width = 40
            
            fila = 2
            fecha_actual = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            for categoria, vars_dict in self.vars.items():
                resultado = vars_dict.get('resultado')
                if resultado and resultado.get():
                    ws[f'A{fila}'] = fecha_actual
                    ws[f'B{fila}'] = categoria
                    ws[f'C{fila}'] = resultado.get()
                    fila += 1
            
            nombre_archivo = f"Resultados_Calculadora_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(nombre_archivo)
            messagebox.showinfo("Éxito", f"Archivo guardado como {nombre_archivo}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al guardar el archivo: {str(e)}")

if __name__ == "__main__":
    app = CalculadoraPinones()
    app.ventana.mainloop()